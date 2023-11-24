import pandas as pd
import numpy as np
import statsmodels.api as sm
from openpyxl import load_workbook,Workbook
import seaborn as sns
import matplotlib.pyplot as plt
if __name__ == '__main__':
    IDP=pd.read_csv('/home/yuling/visual/rfMRI.T1.DWI.SWI.IDPs_ses-01.csv')
    rds=pd.read_csv('/home/yuling/visual/RDS_Scores_2023_Sept_26.csv')
    annot=pd.read_csv('/home/yuling/visual/restIDPs_annot.csv')
    #correlation coefficients for 25 ICA components
    res_25=pd.read_excel('/home/yuling/visual/correlation_25.xlsx')
    #correlation coefficients for 100 ICA components
    res_100=pd.read_excel('/home/yuling/visual/correlation_100.xlsx')
    out_motion=pd.read_excel('/home/yuling/visual/out_motion.xlsx')
    sex_age=pd.read_excel('/home/yuling/visual/age_sex.xlsx')
    network=['Yeo1','Yeo2','Yeo3','Yeo4','Yeo5','Yeo6','Yeo7','subcortical','cerebellum']
    #name of 9 network
    network_list=['visual','somatomotor','dorsal attention','ventral attention','limbic','frontoparietal','default','subcortical','cerebellum']
    #create disctionary for storing network interactions
    namelist={}
    namelist_1={}
    namelist_2={}
    x_net_2=[]
    y_1=[]
    y_2=[]
    y_3=[]
    y_sex=[]
    y_age=[]
    y_motion=[]
    for i in range(0,len(annot)):
        if str(annot.loc[i,'variable']).startswith('F100') :
            
            first=annot.loc[i,'Node1_from_signal']
            second=annot.loc[i,'Node2_from_signal']
            #print(type(first))
            ###first:str
            if int(first) in res_100.columns:
                if int(second) in res_100.columns:
                    
                    corr_1=res_100.loc[20,int(first)]
                    corr_2=res_100.loc[20,int(second)]
                    for j in range(0,len(network)):
                        if corr_1==network[j]:
                            num_1=j
                        if corr_2==network[j]:
                            num_2=j
                    if num_1<=num_2:
                        name=network_list[num_1]+'-'+network_list[num_2]
                        
                    else:
                        name=network_list[num_2]+'-'+network_list[num_1]
                    x_net_2.append(network_list[num_1])
                    
                    x_net_2.append(network_list[num_2])
                    if name in namelist:
                        col_name_100=['eid']
                        col_name_100.append(annot.loc[i,'variable'])
                        data_A100=IDP.loc[:,col_name_100]
                        A25=pd.merge(data_A100,rds)
                        A25=pd.merge(A25,sex_age)
                        #rds_F100=A25.replace([np.inf,-np.inf],np.nan).dropna()
                        A25['age_sex']=A25['age']*A25['sex']
                        data_c=pd.merge(A25,out_motion)
                        #remove missing values
                        data_c=data_c.replace([np.inf,-np.inf],np.nan).dropna()
                        #demean variables
                        data_c[annot.loc[i,'variable']]=(data_c[annot.loc[i,'variable']]-data_c[annot.loc[i,'variable']].mean())/data_c[annot.loc[i,'variable']].std()
                        data_c['assessment_center']=(data_c['assessment_center']-data_c['assessment_center'].mean())/data_c['assessment_center'].std()
                        data_c['head_scaling']=(data_c['head_scaling']-data_c['head_scaling'].mean())/data_c['head_scaling'].std()
                        data_c['head_motion']=(data_c['head_motion']-data_c['head_motion'].mean())/data_c['head_motion'].std()
                        data_c['RDS-2.0']=(data_c['RDS-2.0']-data_c['RDS-2.0'].mean())/data_c['RDS-2.0'].std()
                        data_c['sex']=(data_c['sex']-data_c['sex'].mean())/data_c['sex'].std()
                        data_c['age']=(data_c['age']-data_c['age'].mean())/data_c['age'].std()
                        data_c['age_sex']=(data_c['age_sex']-data_c['age_sex'].mean())/data_c['age_sex'].std()
                        #regression coefficients without covariates
                        reg=sm.OLS(data_c['RDS-2.0'],data_c[annot.loc[i,'variable']])
                        est=reg.fit()
                        namelist[name].append(est.params[0])
                        y_1.append(est.params[0])
                        y_1.append(est.params[0])
                        reg_1=pd.DataFrame(data_c,columns=['assessment_center','head_scaling','head_motion','sex',annot.loc[i,'variable']])
                        reg_m=sm.OLS(data_c['RDS-2.0'],reg_1)
                        est_motion=reg_m.fit()
                        namelist_1[name].append(est_motion.params[4])
                        y_2.append(est_motion.params[4])
                        y_2.append(est_motion.params[4])
                        
                         
                        reg_2=pd.DataFrame(data_c,columns=['assessment_center','head_scaling','head_motion','age','sex','age_sex',annot.loc[i,'variable']])
                        reg_all=sm.OLS(data_c['RDS-2.0'],reg_2)
                        
                        est_c= reg_all.fit()
                        namelist_2[name].append(est_c.params[6])
                        y_3.append(est_c.params[6])
                        y_3.append(est_c.params[6])
                    #######confounds
                        y_sex.append(abs(est_c.params[4]))
                        y_sex.append(abs(est_c.params[4]))

                        y_age.append(abs(est_c.params[3]))
                        y_age.append(abs(est_c.params[3]))

                        y_motion.append(abs(est_c.params[2]))
                        y_motion.append(abs(est_c.params[2]))
                    else: 
                        col_name_100=['eid']
                        col_name_100.append(annot.loc[i,'variable'])
                        data_A100=IDP.loc[:,col_name_100]
                        A25=pd.merge(data_A100,rds)
                        A25=pd.merge(A25,sex_age)
                        #create variable for age_sex interaction
                        A25['age_sex']=A25['age']*A25['sex']
                        
                        data_c=pd.merge(A25,out_motion)
                        data_c=data_c.replace([np.inf,-np.inf],np.nan).dropna()
                        data_c[annot.loc[i,'variable']]=(data_c[annot.loc[i,'variable']]-data_c[annot.loc[i,'variable']].mean())/data_c[annot.loc[i,'variable']].std()
                        data_c['assessment_center']=(data_c['assessment_center']-data_c['assessment_center'].mean())/data_c['assessment_center'].std()
                        data_c['head_scaling']=(data_c['head_scaling']-data_c['head_scaling'].mean())/data_c['head_scaling'].std()
                        data_c['head_motion']=(data_c['head_motion']-data_c['head_motion'].mean())/data_c['head_motion'].std()
                        data_c['RDS-2.0']=(data_c['RDS-2.0']-data_c['RDS-2.0'].mean())/data_c['RDS-2.0'].std()
                        data_c['sex']=(data_c['sex']-data_c['sex'].mean())/data_c['sex'].std()
                        data_c['age']=(data_c['age']-data_c['age'].mean())/data_c['age'].std()
                        data_c['age_sex']=(data_c['age_sex']-data_c['age_sex'].mean())/data_c['age_sex'].std()
                        #regression coefficients without covariates
                        reg=sm.OLS(data_c['RDS-2.0'],data_c[annot.loc[i,'variable']])
                        est=reg.fit()
                        namelist[name]=[est.params[0]]
                        y_1.append(est.params[0])
                        y_1.append(est.params[0])
                        reg_1=pd.DataFrame(data_c,columns=['assessment_center','head_scaling','head_motion','sex',annot.loc[i,'variable']])
                        reg_m=sm.OLS(data_c['RDS-2.0'],reg_1)
                        est_motion=reg_m.fit()
                        namelist_1[name]=[est_motion.params[4]]
                        y_2.append(est_motion.params[4])
                        y_2.append(est_motion.params[4])
                        
                         
                        reg_2=pd.DataFrame(data_c,columns=['assessment_center','head_scaling','head_motion','age','sex','age_sex',annot.loc[i,'variable']])
                        reg_all=sm.OLS(data_c['RDS-2.0'],reg_2)
                        
                        est_c= reg_all.fit()
                        namelist_2[name]=[est_c.params[6]]
                        y_3.append(est_c.params[6])
                        y_3.append(est_c.params[6]) 
                         #######confounds
                        y_sex.append(abs(est_c.params[4]))
                        y_sex.append(abs(est_c.params[4]))

                        y_age.append(abs(est_c.params[3]))
                        y_age.append(abs(est_c.params[3]))

                        y_motion.append(abs(est_c.params[2]))
                        y_motion.append(abs(est_c.params[2]))
        if str(annot.loc[i,'variable']).startswith('F25') :
            first=annot.loc[i,'Node1_from_signal']
            second=annot.loc[i,'Node2_from_signal']
            ###first:str
            if int(first) in res_25.columns:
                if int(second) in res_25.columns:
                    
                    corr_1=res_25.loc[20,int(first)]
                    corr_2=res_25.loc[20,int(second)]
                    for j in range(0,len(network)):
                        if corr_1==network[j]:
                            num_1=j
                        if corr_2==network[j]:
                            num_2=j
                    if num_1<=num_2:
                        name=network_list[num_1]+'-'+network_list[num_2]
                    else:
                        name=network_list[num_2]+'-'+network_list[num_1]
                    x_net_2.append(network_list[num_1])
                    
                    x_net_2.append(network_list[num_2])
                    if name in namelist:
                        col_name_100=['eid']
                        col_name_100.append(annot.loc[i,'variable'])
                        data_A100=IDP.loc[:,col_name_100]
                        A25=pd.merge(data_A100,rds)
                        A25=pd.merge(A25,sex_age)
                        #rds_F100=A25.replace([np.inf,-np.inf],np.nan).dropna()
                        A25['age_sex']=A25['age']*A25['sex']
                         
                        data_c=pd.merge(A25,out_motion)
                        data_c=data_c.replace([np.inf,-np.inf],np.nan).dropna()
                        data_c[annot.loc[i,'variable']]=(data_c[annot.loc[i,'variable']]-data_c[annot.loc[i,'variable']].mean())/data_c[annot.loc[i,'variable']].std()
                        data_c['assessment_center']=(data_c['assessment_center']-data_c['assessment_center'].mean())/data_c['assessment_center'].std()
                        data_c['head_scaling']=(data_c['head_scaling']-data_c['head_scaling'].mean())/data_c['head_scaling'].std()
                        data_c['head_motion']=(data_c['head_motion']-data_c['head_motion'].mean())/data_c['head_motion'].std()
                        data_c['RDS-2.0']=(data_c['RDS-2.0']-data_c['RDS-2.0'].mean())/data_c['RDS-2.0'].std()
                        data_c['sex']=(data_c['sex']-data_c['sex'].mean())/data_c['sex'].std()
                        data_c['age']=(data_c['age']-data_c['age'].mean())/data_c['age'].std()
                        data_c['age_sex']=(data_c['age_sex']-data_c['age_sex'].mean())/data_c['age_sex'].std()
                        #regression coefficients without covariates
                        reg=sm.OLS(data_c['RDS-2.0'],data_c[annot.loc[i,'variable']])
                        est=reg.fit()
                        namelist[name].append(est.params[0])
                        y_1.append(est.params[0])
                        y_1.append(est.params[0])
                        reg_1=pd.DataFrame(data_c,columns=['assessment_center','head_scaling','head_motion','sex',annot.loc[i,'variable']])
                        reg_m=sm.OLS(data_c['RDS-2.0'],reg_1)
                        est_motion=reg_m.fit()
                        namelist_1[name].append(est_motion.params[4])
                        y_2.append(est_motion.params[4])
                        y_2.append(est_motion.params[4])
                        
                         
                        reg_2=pd.DataFrame(data_c,columns=['assessment_center','head_scaling','head_motion','age','sex','age_sex',annot.loc[i,'variable']])
                        reg_all=sm.OLS(data_c['RDS-2.0'],reg_2)
                        
                        est_c= reg_all.fit()
                        namelist_2[name].append(est_c.params[6])
                        y_3.append(est_c.params[6])
                        y_3.append(est_c.params[6])
                         
                          #######confounds
                        y_sex.append(abs(est_c.params[4]))
                        y_sex.append(abs(est_c.params[4]))

                        y_age.append(abs(est_c.params[3]))
                        y_age.append(abs(est_c.params[3]))

                        y_motion.append(abs(est_c.params[2]))
                        y_motion.append(abs(est_c.params[2]))
                    else:   
                        #if network interaction doesn't exist in dictionary
                        col_name_100=['eid']
                        col_name_100.append(annot.loc[i,'variable'])
                        data_A100=IDP.loc[:,col_name_100]
                        A25=pd.merge(data_A100,rds)
                        A25=pd.merge(A25,sex_age)
                        #rds_F100=A25.replace([np.inf,-np.inf],np.nan).dropna()
                        A25['age_sex']=A25['age']*A25['sex']
                         
                        data_c=pd.merge(A25,out_motion)
                        data_c=data_c.replace([np.inf,-np.inf],np.nan).dropna()
                        data_c[annot.loc[i,'variable']]=(data_c[annot.loc[i,'variable']]-data_c[annot.loc[i,'variable']].mean())/data_c[annot.loc[i,'variable']].std()
                        data_c['assessment_center']=(data_c['assessment_center']-data_c['assessment_center'].mean())/data_c['assessment_center'].std()
                        data_c['head_scaling']=(data_c['head_scaling']-data_c['head_scaling'].mean())/data_c['head_scaling'].std()
                        data_c['head_motion']=(data_c['head_motion']-data_c['head_motion'].mean())/data_c['head_motion'].std()
                        data_c['RDS-2.0']=(data_c['RDS-2.0']-data_c['RDS-2.0'].mean())/data_c['RDS-2.0'].std()
                        data_c['sex']=(data_c['sex']-data_c['sex'].mean())/data_c['sex'].std()
                        data_c['age']=(data_c['age']-data_c['age'].mean())/data_c['age'].std()
                        data_c['age_sex']=(data_c['age_sex']-data_c['age_sex'].mean())/data_c['age_sex'].std()
                        #regression coefficients without covariates
                        reg=sm.OLS(data_c['RDS-2.0'],data_c[annot.loc[i,'variable']])
                        est=reg.fit()
                        namelist[name]=[est.params[0]]
                        y_1.append(est.params[0])
                        y_1.append(est.params[0])
                        reg_1=pd.DataFrame(data_c,columns=['assessment_center','head_scaling','head_motion','sex',annot.loc[i,'variable']])
                        reg_m=sm.OLS(data_c['RDS-2.0'],reg_1)
                        est_motion=reg_m.fit()
                        namelist_1[name]=[est_motion.params[4]]
                        y_2.append(est_motion.params[4])
                        y_2.append(est_motion.params[4])
                        
                         
                        reg_2=pd.DataFrame(data_c,columns=['assessment_center','head_scaling','head_motion','age','sex','age_sex',annot.loc[i,'variable']])
                        reg_all=sm.OLS(data_c['RDS-2.0'],reg_2)
                        
                        est_c= reg_all.fit()
                        namelist_2[name]=[est_c.params[6]]
                        y_3.append(est_c.params[6])
                        y_3.append(est_c.params[6])

                        #######confounds
                        y_sex.append(abs(est_c.params[4]))
                        y_sex.append(abs(est_c.params[4]))

                        y_age.append(abs(est_c.params[3]))
                        y_age.append(abs(est_c.params[3]))

                        y_motion.append(abs(est_c.params[2]))
                        y_motion.append(abs(est_c.params[2]))
    '''
    #create a dictionary to store key and mean values for coefficients without covariates
    d={}
    for key,m in namelist.items():
        value_mean=0
        count=0
        for j in m:
            value_mean+=abs(j)
            count+=1
         
        key_mean=value_mean/count
        
        d[key]=[key_mean]
    #order the mean value
    new_d=dict(sorted(d.items(),reverse=True,key=lambda item:item[1]))
    new_x_1=[]
    new_y_1=[]
    new_x_2=[]
    new_y_2=[]
    new_x_age=[]
    new_y_age=[]
    new_x_1_age=[]
    new_y_1_age=[]
    x_tick=[]
    flag_2=0
    for key,value in new_d.items():
        x_tick.append(key)
       
        for m,n in namelist.items():
            if m==key:
                for j in n:
                    new_x.append(flag_2)
                    new_y.append(abs(j))
        for m,n in namelist_c.items():
            if m==key:
                for j in n:
                    new_x_age.append(flag_2+0.4)
                    new_y_age.append(abs(j))
        for m,n in namelist_c_1.items():
            if m==key:
                for j in n:
                    new_x_1_age.append(flag_2+0.2)
                    new_y_1_age.append(abs(j))
        flag_2+=1
    d_age={}
    d_age_1={}
    for key,m in namelist_c.items():
        value_mean=0
        count=0
        for j in m:
            value_mean+=abs(j)
            count+=1
         
        key_mean=value_mean/count
        d_age[key]=[key_mean]
    for key,m in namelist_c_1.items():
        value_mean=0
        count=0
        for j in m:
            value_mean+=abs(j)
            count+=1
         
        key_mean=value_mean/count
        d_age_1[key]=[key_mean]
    #order the coariates for motion
    d_motion={}
    for key,m in namelist_motion.items():
        value_mean=0
        count=0
        for j in m:
            value_mean+=abs(j)
            count+=1
         
        key_mean=value_mean/count
        d_motion[key]=[key_mean]
    #order the mean value
    new_d_motion=[]
    new_d_age=[]
    new_d_age_1=[]
    new_x_c=[]
    new_y_c=[]
    x_t=np.linspace(0,42,43)
    x_t_1=np.linspace(0.3,42.3,43)
    flag_2=0
    for key,value in new_d.items():
        for m,n in namelist_motion.items():
            if m==key:
                for j in n:
                    new_x_c.append(flag_2+0.6)
                    new_y_c.append(abs(j)) 
        flag_2+=1   
        for x,y in d_motion.items():
            if x==key:
                new_d_motion.append(y)
        for x,y in d_age.items():
            if x==key:
                new_d_age.append(y)
        for x,y in d_age_1.items():
            if x==key:
                new_d_age_1.append(y)
    fig,ax=plt.subplots(figsize=(150,25),sharex=False)
    ax.scatter(x=new_x,y=new_y,marker='o',c='r',label='coefficients without covariates')
    ax.legend(loc='upper right',fontsize=25)
    ax.legend(loc='best')  
    
    i=0
    for j in new_d.values():
        if i==0:
            ax.plot([i-0.2,i+0.2],[j,j],c='r',label='mean value for coefficients without covariates')
            i=i+1
        else:
            ax.plot([i-0.2,i+0.2],[j,j],c='r')
            i=i+1
    ax.legend(loc='upper right',fontsize=25)
    ax.scatter(x=new_x_1_age,y=new_y_1_age,marker='o',c='b',label='coefficients with covariates of age and sex')
    ax.legend(loc='upper right',fontsize=25)
    i=0
    for j in new_d_age_1:
        if i==0:
            ax.plot([i+0.2-0.2,i+0.2+0.2],[j,j],c='b',label='mean value for coefficients with age and sex')
            i=i+1
        else:
            ax.plot([i+0.2-0.2,i+0.2+0.2],[j,j],c='b')
            i=i+1
    ax.legend(loc='upper right',fontsize=25)
    #ax[0].set_xticklabels(x_tick)
    #ax.set_ylabel('coefficient')
    ax.scatter(x=new_x_age,y=new_y_age,marker='o',c='y',label='coefficients with covariates of age interated with sex')
    ax.legend(loc='upper right',fontsize=25)
    i=0
    for j in new_d_age:
        if i==0:
            ax.plot([i+0.4-0.2,i+0.4+0.2],[j,j],c='y',label='mean value for coefficients with age interacted with sex')
            i=i+1
        else:
            ax.plot([i+0.4-0.2,i+0.4+0.2],[j,j],c='y')
            i=i+1
    ax.legend(loc='upper right',fontsize=25)
    #ax.set_title('fMRI F100-25-rds coefficients without covariates')
    ####with covariates
    ax.scatter(x=new_x_c,y=new_y_c,marker='o',c='g',label='coefficients with covariates of motion')
    ax.legend(loc='upper right',fontsize=25)
    i=0
    for j in new_d_motion:
        if i==0:
            ax.plot([i+0.6-0.2,i+0.6+0.2],[j,j],c='g',label='mean value for coefficients of motion')
            i=i+1
        else:
            ax.plot([i+0.6-0.2,i+0.6+0.2],[j,j],c='g')
            i=i+1
    ax.legend(loc='upper right',fontsize=25)
    ax.set_xticks(x_t_1,labels=x_tick,rotation=30,fontsize=20)
   
    #plt.xticks(rotation=30)
    ax.set_yticks([0.00,0.01,0.02,0.03,0.04,0.05,0.06,0.07],labels=[0.00,0.01,0.02,0.03,0.04,0.05,0.06,0.07],fontsize=20)
    plt.ylim(0,0.07) 
    plt.xlim(-1,44)
    plt.ylabel('coefficient',fontsize=20)
    plt.xlabel('network',fontsize=20)
    plt.title('fMRI F100-25-rds coefficients with/without covariates',fontsize=40)
    #plt.savefig('fMRI_F100-25-rds_all.png')
    '''
 
    
    #############9 network
    #define the x-axis location
    n_t_1=np.linspace(0.3,8.3,9)
    mean_y_1=[]
    mean_y_2=[]
    mean_y_3=[]
    #order by network: Yeo1-7,subcortical, cerebellum
    new_x_1=[]
    new_y_1=[]
    new_x_2=[]
    new_y_2=[]
    new_x_3=[]
    new_y_3=[]
    new_x_sex=[]
    new_y_sex=[]
    new_x_age=[]
    new_y_age=[]
    new_x_motion=[]
    new_y_motion=[]
    k=0
    for j in network_list:
        amount_1=0
        amount_2=0
        amount_3=0
        #count how many amplitudes each network has
        net_c=0
        for i in range(0,len(x_net_2)):
            if j==x_net_2[i]:
                
                new_x_1.append(k)
                new_y_1.append(y_1[i])
                new_y_sex.append(abs(y_sex[i]))
                amount_1+=y_1[i]
                net_c+=1
                new_x_2.append(k+0.2)
                new_y_2.append(y_2[i])
                new_y_age.append(abs(y_age[i]))
                amount_2+=y_2[i]
                
                new_x_3.append(k+0.4)
                new_y_3.append(y_3[i])
                amount_3+=y_3[i]
                new_y_motion.append(abs(y_motion[i]))
        k+=1
        mean_y_1.append(amount_1/net_c)
        mean_y_2.append(amount_2/net_c)
        mean_y_3.append(amount_3/net_c)
    #visualization
    fig,ax=plt.subplots(figsize=(20,20))    
    ax.scatter(x=new_x_1,y=new_y_1,marker='o',c='r',label='coefficients of IDP without confounds')
    ax.legend(loc='upper right',fontsize=12)
    ax.legend(loc='best')  
    i=0
    for j in mean_y_1:
        if i==0:
            ax.plot([i-0.2,i+0.2],[j,j],c='r',label='mean value for coefficients of IDP without confounds')
            i=i+1
        else:
            ax.plot([i-0.2,i+0.2],[j,j],c='r')
            i=i+1
    ax.legend(loc='upper right',fontsize=12)
   
    #ax.set_title('fMRI A100-25-rds correlation without covariates',fontsize=20)
    #ax.set_xticks([0,1,2,3,4,5,6,7,8],labels=network_list,rotation=30,fontsize=20)
    ax.scatter(x=new_x_2,y=new_y_2,marker='o',c='b',label='coefficients of IDP with confounds except age')
    ax.legend(loc='upper right',fontsize=12)
    i=0
    for j in mean_y_2:
        if i==0:
            ax.plot([i+0.2-0.2,i+0.2+0.2],[j,j],c='b',label='mean value for coefficients of IDP with confounds except age')
            i=i+1
        else:
            ax.plot([i+0.2-0.2,i+0.2+0.2],[j,j],c='b')
            i=i+1
    ax.legend(loc='upper right',fontsize=12)
    
    ax.scatter(x=new_x_3,y=new_y_3,marker='o',c='black',label='coefficients of IDP with all confounds including age_sex interaction')
    ax.legend(loc='upper right',fontsize=12)
    i=0
    for j in mean_y_3:
        if i==0:
            ax.plot([i+0.4-0.2,i+0.4+0.2],[j,j],c='black',label='mean value for coefficients of IDP with all confounds including age_sex interaction')
            i=i+1
        else:
            ax.plot([i+0.4-0.2,i+0.4+0.2],[j,j],c='black')
            i=i+1
    ax.legend(loc='upper right',fontsize=12)

    ax.set_title('fMRI F100-25-rds coefficients of IDP with confounds_9_network',fontsize=30)
    ax.set_yticks([-0.07,-0.06,-0.04,-0.02,0.00,0.02,0.04,0.06,0.08],labels=[-0.07,-0.06,-0.04,-0.02,0.00,0.02,0.04,0.06,0.08],fontsize=20)
    ax.set_xticks(n_t_1,labels=network_list,rotation=30,fontsize=20)
    plt.ylim(-0.07,0.08) 
    plt.xlabel('network',fontsize=25)
    plt.ylabel('coefficient',fontsize=25)
    plt.savefig('F_9netwrok_non_abs.png')     
    ####with covariates
    
                     
    #Amplitude        
    col_name_25=['eid']
    for col in IDP.columns:
        if col.startswith('A25'):
            col_name_25.append(col)
    col_name_100=['eid']
    for col in IDP.columns:
        if col.startswith('A100'):
            col_name_100.append(col)

    data_A25=IDP.loc[:,col_name_25]
    A25=pd.merge(data_A25,rds)
    A25=pd.merge(A25,sex_age)
    A25=pd.merge(A25,out_motion)
    a25_corr=[]
    data_A100=IDP.loc[:,col_name_100]
    A100=pd.merge(data_A100,rds)
    A100=pd.merge(A100,sex_age)
    A100=pd.merge(A100,out_motion)
    a100_corr=[]
     
    #compute regression coefficients
    wb=load_workbook('/home/yuling/visual/correlation_100.xlsx')
    sheet=wb.active
    j=0
    for i in range(1,len(col_name_100)):
        a100_1=A100.loc[:,[col_name_100[i],'RDS-2.0','sex','age','assessment_center','head_scaling','head_motion']].replace([np.inf,-np.inf],np.nan).dropna()
        a100_1['age_sex']=a100_1['age']*a100_1['sex']
        a100_1['RDS-2.0']=(a100_1['RDS-2.0']-a100_1['RDS-2.0'].mean())/a100_1['RDS-2.0'].std()
        a100_1[col_name_100[i]]=(a100_1[col_name_100[i]]-a100_1[col_name_100[i]].mean())/a100_1[col_name_100[i]].std()
        a100_1['assessment_center']=(a100_1['assessment_center']-a100_1['assessment_center'].mean())/a100_1['assessment_center'].std()
        a100_1['head_scaling']=(a100_1['head_scaling']-a100_1['head_scaling'].mean())/a100_1['head_scaling'].std()
        a100_1['head_motion']=(a100_1['head_motion']-a100_1['head_motion'].mean())/a100_1['head_motion'].std()
        a100_1['age']=(a100_1['age']-a100_1['age'].mean())/a100_1['age'].std()
        a100_1['sex']=(a100_1['sex']-a100_1['sex'].mean())/a100_1['sex'].std()
        a100_1['age_sex']=(a100_1['age_sex']-a100_1['age_sex'].mean())/a100_1['age_sex'].std()
        reg=sm.OLS(a100_1['RDS-2.0'],a100_1[col_name_100[i]])
        est=reg.fit()
         
        reg_1=pd.DataFrame(a100_1,columns=['head_motion','head_scaling','assessment_center','sex',col_name_100[i]])
        reg_m=sm.OLS(a100_1['RDS-2.0'],reg_1)
        est_1=reg_m.fit()
        #print(est_1.summary())
        reg_2=pd.DataFrame(a100_1,columns=['head_motion','head_scaling','assessment_center','sex','age','age_sex',col_name_100[i]])
        reg_all=sm.OLS(a100_1['RDS-2.0'],reg_2)
        est_2=reg_all.fit()
        #IDP--confounds
        reg_3=pd.DataFrame(a100_1,columns=['sex'])
        reg_al=sm.OLS(a100_1[col_name_100[i]],reg_3)
        est_3=reg_al.fit()
        reg_4=pd.DataFrame(a100_1,columns=['age'])
        reg_a=sm.OLS(a100_1[col_name_100[i]],reg_4)
        est_4=reg_a.fit()
        reg_5=pd.DataFrame(a100_1,columns=['head_motion'])
        reg_a_1=sm.OLS(a100_1[col_name_100[i]],reg_5)
        est_5=reg_a_1.fit()
        #Age_IDP
        reg_age_idp=sm.OLS(a100_1['age'],a100_1[col_name_100[i]])
        est_age_idp=reg_age_idp.fit()
        #sheet.cell(row=29,column=j+2).value=col_name_100[i]
        sheet.cell(row=30,column=j+2).value=est.params[0]
        sheet.cell(row=31,column=j+2).value=est_1.params[4]
        sheet.cell(row=32,column=j+2).value=est_2.params[6]
        sheet.cell(row=33,column=j+2).value=est_3.params[0]
        sheet.cell(row=34,column=j+2).value=est_4.params[0]
        sheet.cell(row=35,column=j+2).value=est_5.params[0]
        sheet.cell(row=36,column=j+2).value=est_age_idp.params[0]
        j+=1
    wb.save('/home/yuling/visual/correlation_100.xlsx')
    wb=load_workbook('/home/yuling/visual/correlation_25.xlsx')
    sheet=wb.active
    j=0
    for i in range(1,len(col_name_25)):
        a100_1=A25.loc[:,[col_name_25[i],'RDS-2.0','sex','age','assessment_center','head_scaling','head_motion']].replace([np.inf,-np.inf],np.nan).dropna()
        a100_1['age_sex']=a100_1['age']*a100_1['sex']
        a100_1['RDS-2.0']=(a100_1['RDS-2.0']-a100_1['RDS-2.0'].mean())/a100_1['RDS-2.0'].std()
        a100_1[col_name_25[i]]=(a100_1[col_name_25[i]]-a100_1[col_name_25[i]].mean())/a100_1[col_name_25[i]].std()
        a100_1['assessment_center']=(a100_1['assessment_center']-a100_1['assessment_center'].mean())/a100_1['assessment_center'].std()
        a100_1['head_scaling']=(a100_1['head_scaling']-a100_1['head_scaling'].mean())/a100_1['head_scaling'].std()
        a100_1['head_motion']=(a100_1['head_motion']-a100_1['head_motion'].mean())/a100_1['head_motion'].std()
        a100_1['age']=(a100_1['age']-a100_1['age'].mean())/a100_1['age'].std()
        a100_1['sex']=(a100_1['sex']-a100_1['sex'].mean())/a100_1['sex'].std()
        a100_1['age_sex']=(a100_1['age_sex']-a100_1['age_sex'].mean())/a100_1['age_sex'].std()
        reg=sm.OLS(a100_1['RDS-2.0'],a100_1[col_name_25[i]])
        est=reg.fit()
        
        reg_1=pd.DataFrame(a100_1,columns=['head_motion','head_scaling','assessment_center','sex',col_name_25[i]])
        reg_m=sm.OLS(a100_1['RDS-2.0'],reg_1)
        est_1=reg_m.fit()
        reg_2=pd.DataFrame(a100_1,columns=['head_motion','head_scaling','assessment_center','sex','age','age_sex',col_name_25[i]])
        reg_all=sm.OLS(a100_1['RDS-2.0'],reg_2)
        est_2=reg_all.fit()
         
        reg_3=pd.DataFrame(a100_1,columns=['sex'])
        reg_al=sm.OLS(a100_1[col_name_25[i]],reg_3)
        est_3=reg_al.fit()
        reg_4=pd.DataFrame(a100_1,columns=['age'])
        reg_a=sm.OLS(a100_1[col_name_25[i]],reg_4)
        est_4=reg_a.fit()
        reg_5=pd.DataFrame(a100_1,columns=['head_motion'])
        reg_a_1=sm.OLS(a100_1[col_name_25[i]],reg_5)
        est_5=reg_a_1.fit()
        reg_age_idp=sm.OLS(a100_1['age'],a100_1[col_name_25[i]])
        est_age_idp=reg_age_idp.fit()
        #sheet.cell(row=29,column=j+2).value=col_name_100[i]
        sheet.cell(row=30,column=j+2).value=est.params[0]
        sheet.cell(row=31,column=j+2).value=est_1.params[4]
        sheet.cell(row=32,column=j+2).value=est_2.params[6]
        sheet.cell(row=33,column=j+2).value=est_3.params[0]
        sheet.cell(row=34,column=j+2).value=est_4.params[0]
        sheet.cell(row=35,column=j+2).value=est_5.params[0]
        sheet.cell(row=36,column=j+2).value=est_age_idp.params[0] 
        j+=1
    wb.save('/home/yuling/visual/correlation_25.xlsx')
    res_25=pd.read_excel('/home/yuling/visual/correlation_25.xlsx')
    res_100=pd.read_excel('/home/yuling/visual/correlation_100.xlsx')
    x_net=[]
    y_1=[]
    y_2=[]
    y_3=[]
    
    y_sex=[]
    y_age=[]
    y_motion=[]
    y_age_idp=[]
    #order the value of network(x) for visualization
    for j in range(0,len(network)):
        for i in range(0,len(col_name_25)):
            if res_25.iloc[20,i]==network[j]:
                x_net.append(network_list[j])
                y_1.append(res_25.iloc[28,i])
                y_2.append(res_25.iloc[29,i])
                y_3.append(res_25.iloc[30,i])
                y_sex.append(abs(res_25.iloc[31,i]))
                y_age.append(abs(res_25.iloc[32,i]))
                y_motion.append(abs(res_25.iloc[33,i]))
                y_age_idp.append(res_25.iloc[34,i])
    for j in range(0,len(network)):
        
        for i in range(0,len(col_name_100)):
            if res_100.iloc[20,i]==network[j]:
                x_net.append(network_list[j])
                y_1.append(res_100.iloc[28,i])
                y_2.append(res_100.iloc[29,i])
                y_3.append(res_100.iloc[30,i])
                y_sex.append(abs(res_100.iloc[31,i]))
                y_age.append(abs(res_100.iloc[32,i]))
                y_motion.append(abs(res_100.iloc[33,i]))
                y_age_idp.append(res_100.iloc[34,i])
    x_1_re=[]
    y_1_re=[]
    x_2_re=[]
    y_2_re=[]
    x_3_re=[]
    y_3_re=[]
    x_sex_re=[]
    y_sex_re=[]
    x_age_re=[]
    y_age_re=[]
    x_motion_re=[]
    y_motion_re=[]
    y_age_idp_re=[]
    network_m=[]
    network_age=[]
    network_sex=[]
    network_1=[]
    network_2=[]
    network_3=[]
    network_4=[]
    k=0
    for i in range(0,len(network_list)):
        baseline=network_list[i]
        count_1=0
        mean_value_1=0
        mean_value_2=0
        mean_value_3=0
        mean_value_4=0
        mean_value_5=0
        mean_value_6=0
        mean_value_7=0
        for j in range(0,len(y_1)):
            if x_net[j]==baseline:
                print(x_net[j])
                print(y_3[j])
                mean_value_1+=y_1[j]
                mean_value_2+=y_2[j]
                mean_value_3+=y_3[j]
                mean_value_4+=abs(y_sex[j])
                mean_value_5+=abs(y_age[j])
                mean_value_6+=abs(y_motion[j])
                mean_value_7+=y_age_idp[j]
                count_1+=1
                x_1_re.append(k)
                y_1_re.append(y_1[j])
                x_2_re.append(k+0.2)
                y_2_re.append(y_2[j])
                x_3_re.append(k+0.4)
                y_3_re.append(y_3[j])
                y_age_idp_re.append(y_age_idp[j])
                x_sex_re.append(k)
                y_sex_re.append(abs(y_sex[j]))
                x_age_re.append(k+0.2)
                y_age_re.append(abs(y_age[j]))
                x_motion_re.append(k+0.4)
                y_motion_re.append(abs(y_motion[j]))

        k+=1
        
        network_1.append(mean_value_1/count_1)
        network_2.append(mean_value_2/count_1)
        network_3.append(mean_value_3/count_1)
        network_sex.append(mean_value_4/count_1)
        network_age.append(mean_value_5/count_1)
        network_m.append(mean_value_6/count_1)
        network_4.append(mean_value_7/count_1)
    n1=np.linspace(0.3,8.3,9)
    fig,ax=plt.subplots(figsize=(20, 20), dpi=100)
    
    ax.scatter(x=x_1_re,y=y_1_re,marker='o',c='r',label='coefficients of IDP without confounds')
    #ax.scatter(x=x_net_100,y=y_net_100,marker='o',c='r',label='A100')
    ax.legend(loc='upper right',fontsize=12)
    for i in range(0,len(network_1)):
        if i==0:
            ax.plot([i-0.2,i+0.2],[network_1[i],network_1[i]],c='r',label='mean value for coefficients of IDP without confounds')
        else:
            ax.plot([i-0.2,i+0.2],[network_1[i],network_1[i]],c='r')
    ax.legend(loc='upper right',fontsize=12)
    ax.scatter(x=x_2_re,y=y_2_re,marker='o',c='b',label='coefficients of IDP with all confounds in the analysis except age')
    #ax.scatter(x=x_net_100,y=y_net_100,marker='o',c='r',label='A100')
    ax.legend(loc='upper right',fontsize=12)
    for i in range(0,len(network_2)):
        if i==0:
            ax.plot([i+0.2-0.2,i+0.2+0.2],[network_2[i],network_2[i]],c='b',label='mean value for coefficients of IDP with all confounds in the analysis except age')
        else:
            ax.plot([i+0.2-0.2,i+0.2+0.2],[network_2[i],network_2[i]],c='b')
    ax.legend(loc='upper right',fontsize=12)
    ax.scatter(x=x_3_re,y=y_3_re,marker='o',c='black',label='coefficients of Age')
    #ax.scatter(x=x_net_100,y=y_net_100,marker='o',c='r',label='A100')
    ax.legend(loc='upper right',fontsize=12)
    for i in range(0,len(network_3)):
        if i==0:
            ax.plot([i+0.4-0.2,i+0.4+0.2],[network_3[i],network_3[i]],c='black',label='mean value for coefficients of IDP with all confounds in the analysis including age and age_sex interaction')
        else:
            ax.plot([i+0.4-0.2,i+0.4+0.2],[network_3[i],network_3[i]],c='black')
    ax.legend(loc='upper right',fontsize=12)
    #ax.set_yticks([0.00,0.01,0.02,0.03,0.04,0.05,0.06,0.07],labels=[0.00,0.01,0.02,0.03,0.04,0.05,0.06,0.07],fontsize=20)
    ax.set_xticks(n1,labels=network_list,rotation=30,fontsize=20)
    #plt.ylim(0,0.07)
    #plt.xlim(-1,9)
    plt.ylabel('coefficient',fontsize=20)
    plt.xlabel('network',fontsize=20)
    plt.xticks(rotation=30)
    plt.title('fMRI Amplitude-rds coefficient',fontsize=30)
    plt.savefig('rds_age_non_abs.png')
    ########Age_IDP
    fig,ax=plt.subplots(figsize=(20, 20), dpi=100)
    
    ax.scatter(x=x_1_re,y=y_age_idp_re,marker='o',c='r',label='coefficients of Age--Idp')
    #ax.scatter(x=x_net_100,y=y_net_100,marker='o',c='r',label='A100')
    ax.legend(loc='upper right',fontsize=12)
    for i in range(0,len(network_4)):
        if i==0:
            ax.plot([i-0.2,i+0.2],[network_4[i],network_4[i]],c='r',label='mean value')
        else:
            ax.plot([i-0.2,i+0.2],[network_4[i],network_4[i]],c='r')
    ax.legend(loc='upper right',fontsize=12)
    plt.ylabel('coefficient',fontsize=20)
    plt.xlabel('network',fontsize=20)
    plt.xticks(rotation=30)
    plt.title('fMRI Amplitude-rds coefficient',fontsize=30)
    plt.savefig('age_IDP.png')
    #############IDP_confounds
    n_1=np.linspace(0.3,8.3,9)
    fig,ax=plt.subplots(figsize=(20, 20), dpi=100)
    
    ax.scatter(x=x_sex_re,y=y_sex_re,marker='o',c='r',label='coefficients for sex')
    #ax.scatter(x=x_net_100,y=y_net_100,marker='o',c='r',label='A100')
    ax.legend(loc='upper right',fontsize=15)
    for i in range(0,len(network_sex)):
        if i==0:
            ax.plot([i-0.2,i+0.2],[network_sex[i],network_sex[i]],c='r',label='mean value for coefficients for sex')
        else:
            ax.plot([i-0.2,i+0.2],[network_sex[i],network_sex[i]],c='r')
    ax.legend(loc='upper right',fontsize=15)
    
    ax.scatter(x=x_age_re,y=y_age_re,marker='o',c='b',label='coefficients for age')
    #ax.scatter(x=x_net_100,y=y_net_100,marker='o',c='r',label='A100')
    ax.legend(loc='upper right',fontsize=15)
    for i in range(0,len(network_age)):
        if i==0:
            ax.plot([i+0.2-0.2,i+0.2+0.2],[network_age[i],network_age[i]],c='b',label='mean value for coefficients for age')
        else:
            ax.plot([i+0.2-0.2,i+0.2+0.2],[network_age[i],network_age[i]],c='b')
    ax.legend(loc='upper right',fontsize=15)
    ax.scatter(x=x_motion_re,y=y_motion_re,marker='o',c='black',label='coefficients for head motion')
    #ax.scatter(x=x_net_100,y=y_net_100,marker='o',c='r',label='A100')
    ax.legend(loc='upper right',fontsize=15)
    for i in range(0,len(network_m)):
        if i==0:
            ax.plot([i+0.4-0.2,i+0.4+0.2],[network_m[i],network_m[i]],c='black',label='mean value for coefficients for head motion')
        else:
            ax.plot([i+0.4-0.2,i+0.4+0.2],[network_m[i],network_m[i]],c='black')
    ax.legend(loc='upper right',fontsize=15)
    
    #ax.set_yticks([0.00,0.05,0.10,0.15,0.20,0.25,0.30,0.35,0.40,0.45],labels=[0.00,0.05,0.10,0.15,0.20,0.25,0.30,0.35,0.40,0.45],fontsize=20)
    ax.set_xticks(n_1,labels=network_list,rotation=30,fontsize=20)
    #plt.ylim(0.00,0.45)
    #plt.xlim(-1,9)
    plt.ylabel('coefficient',fontsize=20)
    plt.xlabel('network',fontsize=20)
    plt.xticks(rotation=30)
    plt.title('fMRI Amplitude-rds coefficients for confounds',fontsize=30)
    plt.savefig('IDP_confounds.png')
    
    
